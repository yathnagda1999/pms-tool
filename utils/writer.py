"""
Excel file writers - produce bytes for Streamlit download buttons.
"""
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Generic writer
# ---------------------------------------------------------------------------

def to_excel_bytes(df: pd.DataFrame, sheet_name: str = "Sheet1") -> bytes:
    """Write a DataFrame to Excel and return as bytes.

    Args:
        df: DataFrame to write
        sheet_name: target sheet name

    Returns:
        bytes suitable for st.download_button
    """
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet_name, index=False)
    return buf.getvalue()


def write_session_file(session_df: pd.DataFrame) -> bytes:
    """Write session file with light amber highlight on blank CP Code cells.

    Any row where CP Code is empty, NaN, or 'nan' gets an amber fill
    on the CP Code cell only - a visual flag for ops to fill before Part 2.

    Args:
        session_df: session DataFrame from build_session_file()

    Returns:
        bytes suitable for st.download_button
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Session"

    cols = list(session_df.columns)
    cp_col_idx = (cols.index("CP Code") + 1) if "CP Code" in cols else None

    # Header row - bold
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = Font(bold=True)

    # Data rows
    for row_idx, (_, row) in enumerate(session_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(cols, start=1):
            val = row[col_name]
            if pd.isna(val) if not isinstance(val, str) else False:
                val = ""
            ws.cell(row=row_idx, column=col_idx, value=val)

        # Highlight CP Code cell if blank
        if cp_col_idx is not None:
            cp_val = str(row.get("CP Code", "")).strip().lower()
            if not cp_val or cp_val == "nan":
                ws.cell(row=row_idx, column=cp_col_idx).fill = CP_BLANK_FILL

    # Auto-width
    for col_idx, col_name in enumerate(cols, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            max(
                (len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(2, ws.max_row + 1)),
                default=0,
            ),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Allocation file writer - formatted
# ---------------------------------------------------------------------------

CURRENCY_COLS = {
    "InputBrokerage", "InputSTT", "InputStampDuty", "InputSEBIChrg",
    "InputTurnOver", "InputOtherCharges", "InputGST",
    "InputNetAmount", "InputNetRate",
}

# Per-column Excel number format - overrides the default "0.00" where needed
_COL_NUMBER_FORMAT = {
    "InputTurnOver": "0.00",
}

HEADER_FONT    = Font(name="Aptos Narrow", bold=True, size=11)
DATA_FONT      = Font(name="Aptos Narrow", size=11)
CP_BLANK_FILL  = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")  # light amber

_THIN_SIDE     = Side(style="thin")
_CELL_BORDER   = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)


def write_allocation_file(allocation_df: pd.DataFrame) -> bytes:
    """Write the Orbis allocation file with formatting.

    - Header: bold Aptos Narrow 11, no fill, centered, thin border
    - Data: Aptos Narrow 11, centered (Client Name left), thin border
    - Number format '0.00' on all charge and net columns
    - Date format on TradeDate
    - Auto-width columns
    - Settlement No always blank

    Args:
        allocation_df: 19-column allocation DataFrame from allocator

    Returns:
        bytes suitable for st.download_button
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Allocation"

    cols = list(allocation_df.columns)

    # Header row — bold Aptos Narrow, no fill, full border, centered
    for col_idx, col_name in enumerate(cols, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border    = _CELL_BORDER

    # Data rows
    for row_idx, (_, row) in enumerate(allocation_df.iterrows(), start=2):
        for col_idx, col_name in enumerate(cols, start=1):
            val = row[col_name]
            cell = ws.cell(row=row_idx, column=col_idx, value=val)

            cell.font   = DATA_FONT
            cell.border = _CELL_BORDER

            if col_name in CURRENCY_COLS and val is not None:
                cell.number_format = _COL_NUMBER_FORMAT.get(col_name, "0.00")
            elif col_name == "TradeDate" and val is not None:
                cell.number_format = "DD-MM-YYYY"
            elif col_name == "Settlement No":
                cell.value = None  # always blank

            # Alignment: Client Name → left+center; all other columns → center+center
            if col_name == "Client Name":
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")

    # Auto-width - default=0 guards against empty DataFrame (no data rows)
    for col_idx, col_name in enumerate(cols, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(col_name)),
            max(
                (len(str(ws.cell(row=r, column=col_idx).value or ""))
                 for r in range(2, ws.max_row + 1)),
                default=0,
            ),
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 30)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
