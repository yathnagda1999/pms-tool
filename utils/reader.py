"""
Input file readers for all PMS tool formats.
All functions accept a file-like object (BytesIO or UploadedFile).
All column lookups are dynamic - never hardcoded by index.
Both .xls and .xlsx formats are accepted for every file.
"""
from io import BytesIO

import pandas as pd
import xlrd


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _read_file_bytes(file) -> bytes:
    """Read a file-like object or bytes into raw bytes."""
    if hasattr(file, "read"):
        return file.read()
    return file


def _is_xls(content: bytes) -> bool:
    """Return True if content is an OLE2 (.xls) file, False for ZIP (.xlsx)."""
    return content[:4] == b'\xd0\xcf\x11\xe0'


def _get_sheet_rows(content: bytes, sheet_name: str) -> list[tuple]:
    """Return all rows from a named sheet as a list of tuples.

    Auto-detects .xls (xlrd) vs .xlsx (openpyxl) from magic bytes.
    Raises ValueError with helpful message if sheet not found.

    Args:
        content: raw file bytes
        sheet_name: name of the sheet to read

    Returns:
        list of row tuples
    """
    if _is_xls(content):
        wb = xlrd.open_workbook(file_contents=content)
        if sheet_name not in wb.sheet_names():
            raise ValueError(
                f"Could not find sheet '{sheet_name}'. "
                f"Sheets found in this file: {wb.sheet_names()}. "
                f"Please check you uploaded the correct file."
            )
        ws = wb.sheet_by_name(sheet_name)
        return [
            tuple(ws.cell_value(i, j) for j in range(ws.ncols))
            for i in range(ws.nrows)
        ]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Could not find sheet '{sheet_name}'. "
                f"Sheets found in this file: {wb.sheetnames}. "
                f"Please check you uploaded the correct file."
            )
        return list(wb[sheet_name].iter_rows(values_only=True))


def _pd_read_excel(content: bytes, sheet_name: str, **kwargs) -> pd.DataFrame:
    """Read a sheet into a DataFrame using the correct engine for xls/xlsx.

    Args:
        content: raw file bytes
        sheet_name: sheet to read
        **kwargs: passed directly to pd.read_excel

    Returns:
        pd.DataFrame
    """
    engine = "xlrd" if _is_xls(content) else "openpyxl"
    return pd.read_excel(BytesIO(content), sheet_name=sheet_name,
                         engine=engine, **kwargs)


def _normalise_ofin(series: pd.Series) -> pd.Series:
    """Cast OFIN to string, strip whitespace."""
    return series.astype(str).str.strip()


def _require_columns(df: pd.DataFrame, required: list[str], source: str) -> None:
    """Raise ValueError if any required column is missing."""
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"{source}: missing required column(s): {missing}. "
            f"Columns found: {list(df.columns)}"
        )


# ---------------------------------------------------------------------------
# Research Team File
# ---------------------------------------------------------------------------

RESEARCH_REQUIRED = ["S.No", "OFIN", "Client", "Ticker", "Direction",
                     "Qty", "Ref Price", "Value", "CP Code"]

# Accepted aliases for each standard column name (case-insensitive, stripped)
_RESEARCH_ALIASES = {
    "S.No":      ["S.No", "S.NO", "SNO", "SR NO", "SR. NO"],
    "OFIN":      ["OFIN", "OFIN Code", "OFIN CODE"],
    "Client":    ["Client", "Client Name"],
    "Ticker":    ["Ticker", "Stock"],
    "Direction": ["Direction", "Action"],
    "Qty":       ["Qty", "Quantity"],
    "Ref Price": ["Ref Price", "Price"],
    "Value":     ["Value", "Amount"],
    "CP Code":   ["CP Code", "CP CODE"],
}

# Flat uppercase set for quick header-row detection
_RESEARCH_ALIAS_SET = {
    alias.upper()
    for aliases in _RESEARCH_ALIASES.values()
    for alias in aliases
}


def _get_research_sheet_rows(content: bytes) -> list[tuple]:
    """Find and return rows from the research sheet in xls/xlsx content.

    Tries 'Orders' first. If not found, scans all sheets and picks the one
    whose first 10 rows contain the most research column alias matches.
    This handles any sheet name the research team may use.

    Args:
        content: raw file bytes

    Returns:
        list of row tuples from the best-matching sheet

    Raises:
        ValueError: if no sheet with research columns is found
    """
    if _is_xls(content):
        wb = xlrd.open_workbook(file_contents=content)
        sheet_names = wb.sheet_names()
        get_rows = lambda name: [
            tuple(wb.sheet_by_name(name).cell_value(i, j)
                  for j in range(wb.sheet_by_name(name).ncols))
            for i in range(wb.sheet_by_name(name).nrows)
        ]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        sheet_names = wb.sheetnames
        get_rows = lambda name: list(wb[name].iter_rows(values_only=True))

    # Try 'Orders' first
    if "Orders" in sheet_names:
        return get_rows("Orders")

    # Scan all sheets - pick the one with the most alias matches in first 10 rows
    best_sheet, best_score = None, 0
    for name in sheet_names:
        rows = get_rows(name)
        score = 0
        for row in rows[:10]:
            row_vals = [str(v).strip().upper() for v in row if v is not None]
            score = max(score, sum(1 for v in row_vals if v in _RESEARCH_ALIAS_SET))
        if score > best_score:
            best_score, best_sheet = score, name
            best_rows = rows

    if best_sheet is None or best_score < 4:
        raise ValueError(
            f"Research file: could not find a sheet with order data. "
            f"Sheets found: {sheet_names}. "
            f"Expected columns like 'Ticker'/'Stock', 'OFIN'/'OFIN Code', 'Qty', 'Direction'/'Action'."
        )

    return best_rows


def _find_research_header_row(rows: list) -> int:
    """Scan rows top-down to find the header row.

    The header row is the first row where at least 4 cells match known
    research file column aliases. Handles empty rows at the top of the sheet.

    Args:
        rows: all rows from the worksheet

    Returns:
        index of the header row

    Raises:
        ValueError: if no header row found
    """
    for i, row in enumerate(rows):
        row_vals = [str(v).strip().upper() for v in row if v is not None]
        matches = sum(1 for v in row_vals if v in _RESEARCH_ALIAS_SET)
        if matches >= 4:
            return i
    raise ValueError(
        "Research file: could not find a header row. "
        "Expected columns like 'S.No', 'OFIN'/'OFIN Code', "
        "'Ticker'/'Stock', 'Direction'/'Action', 'Qty', 'Price'/'Ref Price'."
    )


def _map_research_headers(raw_headers: list) -> list:
    """Map raw header names to internal standard names via alias dict.

    Strips whitespace and matches case-insensitively. Unknown headers are
    kept as-is (extra columns are harmless).

    Args:
        raw_headers: list of raw cell values from the header row

    Returns:
        list of standardised column names
    """
    result = []
    for h in raw_headers:
        h_clean = str(h).strip() if h is not None else ""
        h_upper = h_clean.upper()
        standard = h_clean  # default: keep as-is
        for std_name, aliases in _RESEARCH_ALIASES.items():
            if h_upper in [a.upper() for a in aliases]:
                standard = std_name
                break
        result.append(standard)
    return result


def read_research_file(file) -> pd.DataFrame:
    """Parse the research team Excel order file.

    Accepts .xls and .xlsx. Handles:
    - Empty rows at the top of the sheet (header row auto-detected)
    - Two column naming formats (e.g. 'Ticker'/'Stock', 'OFIN'/'OFIN Code')
    - Extra whitespace around column headers
    - Any column order

    Args:
        file: BytesIO or Streamlit UploadedFile (.xls or .xlsx)

    Returns:
        pd.DataFrame with columns matching RESEARCH_REQUIRED.
        Direction normalised to uppercase. All strings stripped.
    """
    content = _read_file_bytes(file)
    rows = _get_research_sheet_rows(content)

    if not rows:
        raise ValueError("Research file 'Orders' sheet is empty.")

    # Auto-detect header row (handles empty rows at top)
    header_row_idx = _find_research_header_row(rows)

    # Map raw headers to standard names
    headers = _map_research_headers(rows[header_row_idx])
    data = rows[header_row_idx + 1:]
    df = pd.DataFrame(data, columns=headers)

    # Drop completely empty rows
    df = df.dropna(how="all").reset_index(drop=True)

    _require_columns(df, RESEARCH_REQUIRED, "Research file")

    # Normalise
    df["OFIN"] = _normalise_ofin(df["OFIN"])
    df["Direction"] = df["Direction"].astype(str).str.strip().str.upper()
    df["Ticker"] = df["Ticker"].astype(str).str.strip()
    df["Client"] = df["Client"].astype(str).str.strip()
    df["CP Code"] = df["CP Code"].astype(str).str.strip()
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce")
    df["Ref Price"] = pd.to_numeric(df["Ref Price"], errors="coerce")

    return df


# ---------------------------------------------------------------------------
# Orbis Bank Book
# ---------------------------------------------------------------------------

def read_bank_book(file) -> dict[str, float]:
    """Parse the Orbis Bank Balance Summary sheet.

    Accepts .xls and .xlsx - format auto-detected from magic bytes.

    Args:
        file: BytesIO or Streamlit UploadedFile (.xls or .xlsx)

    Returns:
        dict mapping OFIN (str) -> cash balance (float)
    """
    content = _read_file_bytes(file)
    rows = _get_sheet_rows(content, "Bank Balance Summary")

    # Locate header row dynamically - find row containing "OFIN Code"
    header_row_idx = None
    ofin_col = None
    balance_col = None

    for i, row in enumerate(rows):
        row_vals = [str(v).strip() if v is not None else "" for v in row]
        if "OFIN Code" in row_vals:
            header_row_idx = i
            ofin_col = row_vals.index("OFIN Code")
            balance_col = row_vals.index("Balance")
            break

    if header_row_idx is None:
        raise ValueError(
            "Bank Book: could not find header row containing 'OFIN Code'. "
            "Please check you uploaded the correct file."
        )

    bank_book: dict[str, float] = {}
    current_ofin: str | None = None

    for row in rows[header_row_idx + 1:]:
        row_vals = list(row)

        # Skip "Total" rows
        is_total = any(
            str(v).strip().lower() == "total"
            for v in row_vals
            if v is not None
        )
        if is_total:
            continue

        # Skip empty rows
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        # Data row: read OFIN and balance
        ofin_val = row_vals[ofin_col]
        balance_val = row_vals[balance_col]

        if ofin_val is not None and str(ofin_val).strip():
            current_ofin = str(ofin_val).strip()

        if current_ofin is not None and balance_val is not None:
            try:
                bank_book[current_ofin] = float(balance_val)
            except (ValueError, TypeError):
                pass

    return bank_book


# ---------------------------------------------------------------------------
# Orbis Scrip-wise Report
# ---------------------------------------------------------------------------

def read_scrip_wise_report(file) -> pd.DataFrame:
    """Parse the Orbis Scripwise Clientwise Valuation Report.

    Accepts .xls and .xlsx - format auto-detected from magic bytes.

    Args:
        file: BytesIO or Streamlit UploadedFile (.xls or .xlsx)

    Returns:
        pd.DataFrame with columns [OFIN, Scrip Name, ISIN, Quantity]
    """
    content = _read_file_bytes(file)

    # Sheet name changes daily (first client's name) - always use first sheet
    if _is_xls(content):
        wb = xlrd.open_workbook(file_contents=content)
        ws = wb.sheet_by_index(0)
        rows = [
            tuple(ws.cell_value(i, j) for j in range(ws.ncols))
            for i in range(ws.nrows)
        ]
    else:
        import openpyxl
        wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))

    # Locate header row - find row containing "Scrip Name"
    header_row_idx = None
    col_scrip = col_isin = col_client_code = col_qty = None

    for i, row in enumerate(rows):
        row_vals = [str(v).strip() for v in row]
        if "Scrip Name" in row_vals:
            header_row_idx = i
            col_scrip = row_vals.index("Scrip Name")
            col_isin = row_vals.index("Item No") if "Item No" in row_vals else None
            col_client_code = row_vals.index("Client Code") if "Client Code" in row_vals else None
            col_qty = row_vals.index("Quantity") if "Quantity" in row_vals else None
            break

    if header_row_idx is None:
        raise ValueError(
            "Scrip-wise Report: could not find header row containing 'Scrip Name'."
        )
    if any(c is None for c in [col_isin, col_client_code, col_qty]):
        raise ValueError(
            "Scrip-wise Report: missing one of required columns "
            "'Item No', 'Client Code', 'Quantity'."
        )

    records = []
    for row in rows[header_row_idx + 1:]:
        scrip_val = str(row[col_scrip]).strip()

        # Skip "Scrip Total" rows and empty rows
        if scrip_val.lower() == "scrip total" or not scrip_val:
            continue

        # Skip entirely empty rows
        if all(v is None or str(v).strip() == "" for v in row):
            continue

        isin_val = str(row[col_isin]).strip()
        client_code_val = str(row[col_client_code]).strip()
        qty_val = row[col_qty]

        try:
            qty_float = float(qty_val)
        except (ValueError, TypeError):
            qty_float = 0.0

        records.append({
            "OFIN": client_code_val,
            "Scrip Name": scrip_val,
            "ISIN": isin_val,
            "Quantity": qty_float,
        })

    df = pd.DataFrame(records)
    df["OFIN"] = _normalise_ofin(df["OFIN"])
    df["Scrip Name"] = df["Scrip Name"].str.strip()
    return df


# ---------------------------------------------------------------------------
# Session File
# ---------------------------------------------------------------------------

SESSION_REQUIRED = ["S.No", "Batch", "OFIN", "Client", "Ticker",
                    "ISIN", "Direction", "Qty", "Ref Price", "CP Code"]


def read_session_file(file) -> pd.DataFrame:
    """Read an existing session file.

    Accepts .xls and .xlsx - format auto-detected from magic bytes.

    Args:
        file: BytesIO or Streamlit UploadedFile (.xls or .xlsx)

    Returns:
        pd.DataFrame with SESSION_REQUIRED columns
    """
    content = _read_file_bytes(file)
    df = _pd_read_excel(content, sheet_name=0, dtype=str)
    df.columns = [c.strip() for c in df.columns]
    _require_columns(df, SESSION_REQUIRED, "Session file")
    df["OFIN"] = _normalise_ofin(df["OFIN"])
    df["Qty"] = pd.to_numeric(df["Qty"], errors="coerce")
    df["Ref Price"] = pd.to_numeric(df["Ref Price"], errors="coerce")
    df["Batch"] = pd.to_numeric(df["Batch"], errors="coerce").astype("Int64")
    df["S.No"]  = pd.to_numeric(df["S.No"],  errors="coerce").astype("Int64")
    return df


# ---------------------------------------------------------------------------
# Broker Reply - Ambit
# ---------------------------------------------------------------------------

AMBIT_REQUIRED = ["Transaction Date", "Exchange", "ISIN No.", "Transaction Type",
                  "quantity", "Brokerage", "stt", "Stamp Duty", "SEBI Charges",
                  "Turnover Tax", "Other Charges", "GST Amount", "Net Amount"]


def read_broker_reply_ambit(file) -> pd.DataFrame:
    """Parse Ambit broker execution reply.

    Accepts .xls and .xlsx - format auto-detected from magic bytes.

    Args:
        file: BytesIO or Streamlit UploadedFile (.xls or .xlsx)

    Returns:
        pd.DataFrame with normalised column names
    """
    content = _read_file_bytes(file)
    # Validate sheet exists before reading into DataFrame
    _get_sheet_rows(content, "Sheet1")
    df = _pd_read_excel(content, sheet_name="Sheet1")
    df.columns = [str(c).strip() for c in df.columns]
    _require_columns(df, AMBIT_REQUIRED, "Ambit broker reply")
    return df


# ---------------------------------------------------------------------------
# Broker Reply - InCred
# ---------------------------------------------------------------------------

INCRED_SHEET = "Incred_Capital_Trade_Confirmati"
INCRED_REQUIRED = ["Exchange", "ISIN No.", "Transaction Type", "Quantity",
                   "Amount", "Brokerage", "STT", "Stamp Duty", "SEBI Charges",
                   "Turnover Tax", "Other Charges", "GST Amount", "Net Amount"]


def read_broker_reply_incred(file) -> pd.DataFrame:
    """Parse InCred broker execution reply.

    Accepts .xls and .xlsx - format auto-detected from magic bytes.
    Handles string-typed numeric columns and empty GST Amount.

    Args:
        file: BytesIO or Streamlit UploadedFile (.xls or .xlsx)

    Returns:
        pd.DataFrame with all numeric columns cast to float
    """
    content = _read_file_bytes(file)
    # Validate sheet exists before reading into DataFrame
    _get_sheet_rows(content, INCRED_SHEET)
    df = _pd_read_excel(content, sheet_name=INCRED_SHEET, dtype=str)
    df.columns = [str(c).strip() for c in df.columns]
    _require_columns(df, INCRED_REQUIRED, "InCred broker reply")

    # Cast string columns to float
    str_to_float_cols = ["Amount", "Stamp Duty", "SEBI Charges",
                         "Turnover Tax", "Other Charges"]
    for col in str_to_float_cols:
        df[col] = pd.to_numeric(df[col].str.strip(), errors="coerce").fillna(0.0)

    # GST Amount: empty string -> 0.0
    df["GST Amount"] = pd.to_numeric(
        df["GST Amount"].astype(str).str.strip().replace("", "0"),
        errors="coerce"
    ).fillna(0.0)

    # Cast remaining numeric cols
    for col in ["Quantity", "Brokerage", "STT", "Net Amount"]:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    return df
